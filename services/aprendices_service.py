"""
services/aprendices_service.py
CRUD para Aprendices + carga masiva desde Excel/CSV
"""
import io
import datetime
from werkzeug.security import generate_password_hash
from app import db
from models.models import Aprendiz, Usuario, Rol, Ficha


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _crear_usuario_aprendiz(tipo_doc, ident, correo, pnombre, snombre, papellido, sapellido, contrasena) -> Usuario:
    rol_apr = Rol.query.filter_by(nombre="Aprendiz").first()
    if not rol_apr:
        raise ValueError("Rol 'Aprendiz' no encontrado.")
    u = Usuario(
        tipo_documento=tipo_doc,
        identificacion=ident.strip(),
        correo=correo.strip().lower(),
        primer_nombre=pnombre.strip().upper(),
        segundo_nombre=snombre.strip().upper() if snombre else None,
        primer_apellido=papellido.strip().upper(),
        segundo_apellido=sapellido.strip().upper() if sapellido else None,
        contrasena_hash=generate_password_hash(contrasena),
        rol_id=rol_apr.rol_id,
        cuenta_activa=True,
    )
    return u


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------

def listar_aprendices(ficha_id: str | None = None, estado: str | None = None):
    q = Aprendiz.query.join(Usuario).order_by(Usuario.primer_apellido)
    if ficha_id:
        q = q.filter(Aprendiz.ficha_id == ficha_id)
    if estado:
        q = q.filter(Aprendiz.estado == estado)
    return q.all()


def obtener_aprendiz(aprendiz_id: str):
    return Aprendiz.query.get_or_404(aprendiz_id)


def crear_aprendiz(
    ficha_id: str,
    tipo_documento: str,
    identificacion: str,
    correo: str,
    primer_nombre: str,
    segundo_nombre: str | None,
    primer_apellido: str,
    segundo_apellido: str | None,
    contrasena: str,
    fecha_ingreso: datetime.datetime,
) -> Aprendiz:
    # Reusar usuario existente si ya existe
    usuario = Usuario.query.filter_by(identificacion=identificacion.strip()).first()
    if not usuario:
        usuario = _crear_usuario_aprendiz(
            tipo_documento, identificacion, correo,
            primer_nombre, segundo_nombre,
            primer_apellido, segundo_apellido, contrasena,
        )
        db.session.add(usuario)
        db.session.flush()

    aprendiz = Aprendiz(
        usuario_id=usuario.usuario_id,
        ficha_id=ficha_id,
        estado="Activo",
        fecha_ingreso=fecha_ingreso,
    )
    db.session.add(aprendiz)
    db.session.commit()
    return aprendiz


def editar_aprendiz(aprendiz_id: str, estado: str) -> Aprendiz:
    a = obtener_aprendiz(aprendiz_id)
    a.estado = estado
    if estado == "Retirado":
        a.fecha_retiro = datetime.datetime.utcnow()
    a.fecha_actualizacion = datetime.datetime.utcnow()
    db.session.commit()
    return a


def eliminar_aprendiz(aprendiz_id: str):
    a = obtener_aprendiz(aprendiz_id)
    db.session.delete(a)
    db.session.commit()


# ---------------------------------------------------------------------------
# Carga masiva
# ---------------------------------------------------------------------------

def procesar_carga_masiva(archivo, ficha_id: str, contrasena: str, fecha_ingreso: datetime.datetime) -> dict:
    """
    Procesa un archivo Excel o CSV con aprendices y los crea.
    Retorna dict con: creados, omitidos, errores
    """
    import openpyxl
    import csv

    nombre_archivo = archivo.filename.lower()
    filas = []

    if nombre_archivo.endswith(('.xlsx', '.xls')):
        wb = openpyxl.load_workbook(archivo.stream, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                filas.append([str(c).strip() if c is not None else '' for c in row])
    elif nombre_archivo.endswith('.csv'):
        contenido = archivo.stream.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(contenido))
        next(reader, None)  # saltar encabezado
        for row in reader:
            if any(c.strip() for c in row):
                filas.append([c.strip() for c in row])
    else:
        raise ValueError("Formato de archivo no soportado.")

    creados = 0
    omitidos = 0
    errores = []

    for num_fila, fila in enumerate(filas, start=2):
        try:
            tipo_doc = fila[0] if len(fila) > 0 else ''
            ident = fila[1] if len(fila) > 1 else ''
            pnombre = fila[2] if len(fila) > 2 else ''
            snombre = fila[3] if len(fila) > 3 else ''
            papellido = fila[4] if len(fila) > 4 else ''
            sapellido = fila[5] if len(fila) > 5 else ''
            correo = fila[6] if len(fila) > 6 else ''

            if not all([tipo_doc, ident, pnombre, papellido, correo]):
                errores.append({"fila": num_fila, "mensaje": "Campos obligatorios faltantes."})
                continue

            if tipo_doc not in ('CC', 'TI', 'CE', 'PEP', 'PPT'):
                errores.append({"fila": num_fila, "mensaje": f"Tipo de documento '{tipo_doc}' inválido."})
                continue

            # Verificar si ya está en la ficha
            existente = (
                Aprendiz.query
                .join(Usuario)
                .filter(Usuario.identificacion == ident, Aprendiz.ficha_id == ficha_id)
                .first()
            )
            if existente:
                omitidos += 1
                continue

            crear_aprendiz(
                ficha_id=ficha_id,
                tipo_documento=tipo_doc,
                identificacion=ident,
                correo=correo,
                primer_nombre=pnombre,
                segundo_nombre=snombre or None,
                primer_apellido=papellido,
                segundo_apellido=sapellido or None,
                contrasena=contrasena,
                fecha_ingreso=fecha_ingreso,
            )
            creados += 1

        except Exception as exc:
            db.session.rollback()
            errores.append({"fila": num_fila, "mensaje": str(exc)})

    return {"creados": creados, "omitidos": omitidos, "errores": errores}


def generar_plantilla_excel() -> bytes:
    """Genera una plantilla Excel para carga masiva."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aprendices"
    encabezados = [
        "tipo_documento", "identificacion", "primer_nombre", "segundo_nombre",
        "primer_apellido", "segundo_apellido", "correo"
    ]
    ws.append(encabezados)
    ws.append(["CC", "1234567890", "JUAN", "CARLOS", "PEREZ", "GOMEZ", "juan@correo.com"])
    ws.append(["TI", "987654321", "MARIA", "", "GARCIA", "", "maria@correo.com"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
